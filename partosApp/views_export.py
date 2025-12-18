import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Prefetch
from datetime import datetime
from openpyxl import Workbook

# Imports de modelos
from partosApp.models import RegistroParto
from recienNacidoApp.models import RegistroRecienNacido, DocumentosParto
from ingresoPartoApp.models import FichaParto
from matronaApp.models import FichaObstetrica
from django.contrib.auth.models import User
from django.db.models import Q

@login_required
def vista_exportar_libro(request):
    """Renderiza la vista de selección de datos para exportar"""
    from partosApp.models import CatalogoTipoParto, CatalogoClasificacionRobson
    
    context = {
        'tipos_parto': CatalogoTipoParto.objects.filter(activo=True).order_by('descripcion'),
        'clasificaciones_robson': CatalogoClasificacionRobson.objects.filter(activo=True).order_by('numero_grupo'),
    }
    
    return render(request, 'Partos/exportar_seleccion.html', context)


# ==========================================
# HELPERS
# ==========================================
def safe_get(obj, attr_path, default="-"):
    """Navega seguro por atributos anidados: safe_get(p, 'ficha.paciente.nombre')"""
    if obj is None:
        return default
    try:
        val = obj
        for attr in attr_path.split('.'):
            val = getattr(val, attr)
            if val is None:
                return default
        return str(val)
    except AttributeError:
        return default

def safe_val(obj, attr_path, default=None):
    """Retorna el valor crudo (sin convertir a string) o default"""
    if obj is None:
        return default
    try:
        val = obj
        for attr in attr_path.split('.'):
            val = getattr(val, attr)
            if val is None:
                return default
        return val
    except AttributeError:
        return default

def get_si_no(bool_val):
    return "SI" if bool_val else "NO"


def obt_ficha_parto_safe(parto):
    """Helper para obtener ficha de parto (ingreso) de forma segura"""
    try:
        return parto.ficha_obstetrica.ficha_parto
    except:
        return None

def check_patologia(parto, patologia_str):
    """Verifica en M2M de FichaObstetrica"""
    try:
        patologias = parto.ficha_obstetrica.patologias.all()
        return "SI" if any(p.nombre.upper() in patologia_str.upper() for p in patologias) else "NO"
    except:
        return "NO"

def get_paridad(parto):
    try:
        gestas = parto.ficha_obstetrica.numero_gestas
        partos = parto.ficha_obstetrica.numero_partos
        return f"G{gestas} P{partos}"
    except:
        return "-"

@login_required
def generar_excel_libro(request):
    """
    Genera el archivo Excel 'Libro de Partos' con headers agrupados (2 filas).
    Itera sobre partos y sus recién nacidos (1 fila por RN).
    """
    
    # Configuración de Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Libro_Partos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'

    # Crear Libro y Hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro de Partos"

    # Estilos Base
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_style = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Definición de Colores por Categoría (Super Header)
    # Colores visualmente distintos para separar categorías
    CATEGORY_COLORS = {
        "DATOS GENERALES": "0070C0",       # Azul
        "TOMA DE VIH / EXÁMENES": "E36C09", # Naranja
        "SGB": "7030A0",                   # Púrpura
        "VDRL": "00B050",                  # Verde
        "TRABAJO DE PARTO": "00B0F0",      # Azul claro
        "PARTO": "FF0000",                 # Rojo brillante
        "ANALGESIA": "9966FF",             # Púrpura claro
        "PUERPERIO": "FFC000",             # Amarillo/Dorado
        "RESULTADO DEL RN": "92D050",      # Verde lima
        "APEGO Y ACOMPAÑAMIENTO": "FF33CC", # Rosa
        "INFORMACION DE LOS PROFESIONALES": "7F7F7F", # Gris
        "Ley N° 21.372 Dominga": "993300", # Marrón
        "Placenta": "3399FF",              # Azul cielo
        "Registro Civil": "002060",        # Azul muy oscuro
        "": "000000"                       # Negro (fallback)
    }

    # ---------------------------------------------------------
    # DEFINICIÓN DE COLUMNAS (SuperHeader, Header, Extractor)
    # ---------------------------------------------------------
    # x es un diccionario {'parto': parto, 'rn': rn}
    
    columns_map = [
        # --- DATOS GENERALES (Identificación + Patologías) ---
        ("DATOS GENERALES", "Tipo de Paciente", lambda x: safe_get(x['parto'], 'ficha_obstetrica.tipo_paciente.nombre', 'SIN PREVISIÓN')),
        ("DATOS GENERALES", "Origen Ingreso", lambda x: safe_get(x['parto'], 'ficha_obstetrica.tipo_ingreso', '-')),
        ("DATOS GENERALES", "FECHA", lambda x: (safe_val(x['parto'], 'fecha_hora_parto') or safe_val(x['parto'], 'ficha_obstetrica.fecha_creacion')).strftime('%d/%m/%Y') if (safe_val(x['parto'], 'fecha_hora_parto') or safe_val(x['parto'], 'ficha_obstetrica.fecha_creacion')) else '-'),
        ("DATOS GENERALES", "Hora", lambda x: (safe_val(x['parto'], 'fecha_hora_parto') or safe_val(x['parto'], 'ficha_obstetrica.fecha_creacion')).strftime('%H:%M') if (safe_val(x['parto'], 'fecha_hora_parto') or safe_val(x['parto'], 'ficha_obstetrica.fecha_creacion')) else '-'),
        ("DATOS GENERALES", "PLAN DE PARTO (SI/NO)", lambda x: get_si_no(safe_get(x['parto'], 'ficha_obstetrica.plan_de_parto', False))),
        ("DATOS GENERALES", "Visita Guiada", lambda x: get_si_no(safe_get(x['parto'], 'ficha_obstetrica.visita_guiada', False))),
        ("DATOS GENERALES", "Nombre completo", lambda x: safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.nombre_completo', '-')),
        ("DATOS GENERALES", "RUN", lambda x: safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Rut', '-').split('-')[0] if '-' in safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Rut', '') else safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Rut', '-')),
        ("DATOS GENERALES", "DV", lambda x: safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Rut', '-').split('-')[1] if '-' in safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Rut', '') else '-'),
        ("DATOS GENERALES", "Inmigrante", lambda x: get_si_no(safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Inmigrante', False))),
        ("DATOS GENERALES", "Nacionalidad", lambda x: safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Nacionalidad.nombre', 'CHILENA')),
        ("DATOS GENERALES", "Pueblo Originario", lambda x: safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Pueblos_originarios.nombre', '-')),
        ("DATOS GENERALES", "Edad", lambda x: safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.edad', '-')),
        ("DATOS GENERALES", "Discapacidad", lambda x: get_si_no(safe_get(x['parto'], 'ficha_obstetrica.tiene_discapacidad', False))),
        ("DATOS GENERALES", "Privada de Libertad", lambda x: get_si_no(safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Privada_de_Libertad', False))),
        ("DATOS GENERALES", "Trans Masculino", lambda x: get_si_no(safe_get(x['parto'], 'ficha_obstetrica.paciente.persona.Trans_Masculino', False))),
        ("DATOS GENERALES", "IMC", lambda x: safe_get(x['parto'], 'ficha_obstetrica.imc', '-')),
        ("DATOS GENERALES", "Peso Actual (kg)", lambda x: safe_get(x['parto'], 'ficha_obstetrica.peso_actual', '-')),
        ("DATOS GENERALES", "Talla Actual (cm)", lambda x: safe_get(x['parto'], 'ficha_obstetrica.talla_actual', '-')),
        ("DATOS GENERALES", "Tipo de Discapacidad", lambda x: safe_get(x['parto'], 'ficha_obstetrica.discapacidad.nombre', '-')),
        ("DATOS GENERALES", "Paridad", lambda x: f"G{safe_get(x['parto'], 'ficha_obstetrica.numero_gestas', '?')} P{safe_get(x['parto'], 'ficha_obstetrica.numero_partos', '?')}"),
        ("DATOS GENERALES", "Número de Abortos", lambda x: safe_get(x['parto'], 'ficha_obstetrica.numero_abortos', '-')),
        ("DATOS GENERALES", "Nacidos Vivos", lambda x: safe_get(x['parto'], 'ficha_obstetrica.nacidos_vivos', '-')),
        ("DATOS GENERALES", "Partos Vaginales", lambda x: safe_get(x['parto'], 'ficha_obstetrica.partos_vaginales', '-')),
        ("DATOS GENERALES", "Cesáreas", lambda x: safe_get(x['parto'], 'ficha_obstetrica.partos_cesareas', '-')),
        ("DATOS GENERALES", "FUM", lambda x: safe_val(x['parto'], 'ficha_obstetrica.fecha_ultima_regla').strftime('%d/%m/%Y') if safe_val(x['parto'], 'ficha_obstetrica.fecha_ultima_regla') else '-'),
        ("DATOS GENERALES", "FPP", lambda x: safe_val(x['parto'], 'ficha_obstetrica.fecha_probable_parto').strftime('%d/%m/%Y') if safe_val(x['parto'], 'ficha_obstetrica.fecha_probable_parto') else '-'),
        ("DATOS GENERALES", "Control Prenatal", lambda x: safe_get(x['parto'], 'ficha_obstetrica.numero_controles', '0')),
        ("DATOS GENERALES", "Consultorio de Origen", lambda x: safe_get(x['parto'], 'ficha_obstetrica.consultorio_origen.nombre', '-')),
        
        # Acompañante
        ("DATOS GENERALES", "Tiene Acompañante", lambda x: get_si_no(safe_get(x['parto'], 'ficha_obstetrica.tiene_acompanante', False))),
        ("DATOS GENERALES", "Nombre Acompañante", lambda x: safe_get(x['parto'], 'ficha_obstetrica.nombre_acompanante', '-')),
        ("DATOS GENERALES", "RUT Acompañante", lambda x: safe_get(x['parto'], 'ficha_obstetrica.rut_acompanante', '-')),
        ("DATOS GENERALES", "Parentesco Acompañante", lambda x: safe_get(x['parto'], 'ficha_obstetrica.parentesco_acompanante', '-')),
        ("DATOS GENERALES", "Teléfono Acompañante", lambda x: safe_get(x['parto'], 'ficha_obstetrica.telefono_acompanante', '-')),
        
        # Contacto de Emergencia
        ("DATOS GENERALES", "Contacto Emergencia - Nombre", lambda x: safe_get(x['parto'], 'ficha_obstetrica.nombre_contacto_emergencia', '-')),
        ("DATOS GENERALES", "Contacto Emergencia - Teléfono", lambda x: safe_get(x['parto'], 'ficha_obstetrica.telefono_emergencia', '-')),
        ("DATOS GENERALES", "Contacto Emergencia - Parentesco", lambda x: safe_get(x['parto'], 'ficha_obstetrica.parentesco_contacto_emergencia', '-')),
        
        # Patologías
        ("DATOS GENERALES", "Preeclampsia Severa", lambda x: get_si_no(safe_get(x['parto'], 'ficha_ingreso_parto.preeclampsia_severa', False))),
        ("DATOS GENERALES", "Eclampsia", lambda x: get_si_no(safe_get(x['parto'], 'ficha_ingreso_parto.eclampsia', False))),
        ("DATOS GENERALES", "Sepsis o Infecc. Grave", lambda x: get_si_no(safe_get(x['parto'], 'ficha_ingreso_parto.sepsis_infeccion_grave', False))),
        ("DATOS GENERALES", "Infección Ovular", lambda x: get_si_no(safe_get(x['parto'], 'ficha_ingreso_parto.infeccion_ovular', False))),

        # --- TOMA DE VIH / EXÁMENES ---
        ("TOMA DE VIH / EXÁMENES", "N° ARO", lambda x: safe_get(x['parto'], 'ficha_obstetrica.clasificacion_aro.nombre', '-')), 
        # Desglose de 3 VIH con fechas
        ("TOMA DE VIH / EXÁMENES", "VIH 1 - Fecha", lambda x: safe_val(x['parto'], 'ficha_obstetrica.vih_1_fecha').strftime('%d/%m/%Y') if safe_val(x['parto'], 'ficha_obstetrica.vih_1_fecha') else '-'),
        ("TOMA DE VIH / EXÁMENES", "VIH 1 - Resultado", lambda x: safe_get(x['parto'], 'ficha_obstetrica.vih_1_resultado', '-')),
        ("TOMA DE VIH / EXÁMENES", "VIH 2 - Fecha", lambda x: safe_val(x['parto'], 'ficha_obstetrica.vih_2_fecha').strftime('%d/%m/%Y') if safe_val(x['parto'], 'ficha_obstetrica.vih_2_fecha') else '-'),
        ("TOMA DE VIH / EXÁMENES", "VIH 2 - Resultado", lambda x: safe_get(x['parto'], 'ficha_obstetrica.vih_2_resultado', '-')),
        ("TOMA DE VIH / EXÁMENES", "VIH 3 (Intraparto)", lambda x: safe_get(x['parto'], 'ficha_ingreso_parto.vih_resultado', '-')),
        
        ("SGB", "Pesquisa", lambda x: get_si_no(safe_get(x['parto'], 'ficha_ingreso_parto.sgb_pesquisa', False))),
        ("SGB", "Resultado", lambda x: safe_get(x['parto'], 'ficha_ingreso_parto.sgb_resultado', '-')),
        ("SGB", "Antibiótico por SGB", lambda x: get_si_no(safe_get(x['parto'], 'ficha_ingreso_parto.antibiotico_sgb', False))),
        ("VDRL", "Resultado VDRL embarazo", lambda x: safe_get(x['parto'], 'ficha_ingreso_parto.vdrl_resultado', '-')),
        ("VDRL", "Tratamiento ATB por Sífilis", lambda x: get_si_no(safe_get(x['parto'], 'ficha_ingreso_parto.tratamiento_sifilis', False))),

        # --- TRABAJO DE PARTO ---
        ("TRABAJO DE PARTO", "Sem. Obst. (semanas)", lambda x: safe_get(x['parto'], 'edad_gestacional_semanas', '-')),
        ("TRABAJO DE PARTO", "Sem. Obst. (días)", lambda x: safe_get(x['parto'], 'edad_gestacional_dias', '-')),
        ("TRABAJO DE PARTO", "Monitor", lambda x: get_si_no(safe_get(x['parto'], 'monitor_continuo', False))),
        ("TRABAJO DE PARTO", "TTC", lambda x: safe_get(x['parto'], 'tiempo_trabajo_parto_total_minutos', '-')),
        ("TRABAJO DE PARTO", "Inducción", lambda x: get_si_no(safe_get(x['parto'], 'induccion', False))),
        ("TRABAJO DE PARTO", "Aceleración ó Corrección", lambda x: get_si_no(safe_get(x['parto'], 'aceleracion', False))),
        ("TRABAJO DE PARTO", "Rotura membrana", lambda x: "SI" if safe_get(x['parto'], 'tipo_rotura_membrana') else "NO"),
        ("TRABAJO DE PARTO", "Tiempo dilatación (min)", lambda x: safe_get(x['parto'], 'tiempo_dilatacion_minutos', '-')),
        ("TRABAJO DE PARTO", "Tiempo Expulsivo (min)", lambda x: safe_get(x['parto'], 'tiempo_expulsivo_minutos', '-')),


        # --- PARTO ---
        ("PARTO", "Libertad de Movimiento", lambda x: get_si_no(safe_get(x['parto'], 'libertad_movimiento', False))),
        ("PARTO", "Tipo de Régimen", lambda x: safe_get(x['parto'], 'regimen_parto.descripcion', '-')),
        ("PARTO", "Tipo de parto", lambda x: safe_get(x['parto'], 'tipo_parto.descripcion', '-')),
        ("PARTO", "Alumbramiento Dirigido", lambda x: get_si_no(safe_get(x['parto'], 'alumbramiento_dirigido', False))),
        ("PARTO", "Clasificación de Robson", lambda x: safe_get(x['parto'], 'clasificacion_robson.descripcion', '-')),
        ("PARTO", "Posición materna parto", lambda x: safe_get(x['parto'], 'posicion_parto.descripcion', '-')),

        # --- ANALGESIA Y ANESTESIA ---
        ("ANALGESIA", "Analgesia No Farmacológica", lambda x: get_si_no(safe_get(x['parto'], 'analgesia_no_farmacologica', False))),
        ("ANALGESIA", "Peridural - Solicitada Paciente", lambda x: get_si_no(safe_get(x['parto'], 'peridur al_solicitada_paciente', False))),
        ("ANALGESIA", "Peridural - Indicada Médico", lambda x: get_si_no(safe_get(x['parto'], 'peridural_indicada_medico', False))),
        ("ANALGESIA", "Peridural - Administrada", lambda x: get_si_no(safe_get(x['parto'], 'peridural_administrada', False))),
        ("ANALGESIA", "Tiempo Espera Peridural (min)", lambda x: safe_get(x['parto'], 'tiempo_espera_peridural_minutos', '-')),
        ("ANALGESIA", "Anestesia Local (Periné)", lambda x: get_si_no(safe_get(x['parto'], 'anestesia_local', False))),
        ("ANALGESIA", "Anestesia General", lambda x: get_si_no(safe_get(x['parto'], 'anestesia_general', False))),
        ("ANALGESIA", "Anestesia Raquídea", lambda x: get_si_no(safe_get(x['parto'], 'anestesia_raquidea', False))),

        # --- PUERPERIO ---
        ("PUERPERIO", "Ofrec. posiciones alt.", lambda x: get_si_no(safe_get(x['parto'], 'ofrecimiento_posiciones_alternativas', False))),
        ("PUERPERIO", "Estado periné", lambda x: safe_get(x['parto'], 'estado_perine.descripcion', '-')),
        ("PUERPERIO", "Esteriliz.", lambda x: get_si_no(safe_get(x['parto'], 'esterilizacion', False))),
        ("PUERPERIO", "Inercia Uterina", lambda x: get_si_no(safe_get(x['parto'], 'inercia_uterina', False))),
        ("PUERPERIO", "Restos Placentarios", lambda x: get_si_no(safe_get(x['parto'], 'restos_placentarios', False))),
        ("PUERPERIO", "Desgarro/Trauma", lambda x: get_si_no(safe_get(x['parto'], 'desgarro_vaginal', False))),
        ("PUERPERIO", "Hemorragia Postparto", lambda x: get_si_no(safe_get(x['parto'], 'hemorragia_postparto', False))),
        ("PUERPERIO", "Tipo Esterilización", lambda x: safe_get(x['parto'], 'tipo_esterilizacion.descripcion', '-')),

        # --- RESULTADO DEL RN (New Image) ---
        ("RESULTADO DEL RN", "Sexo", lambda x: safe_get(x['rn'], 'sexo.descripcion', '-') if x['rn'] else '-'),
        ("RESULTADO DEL RN", "Peso", lambda x: safe_get(x['rn'], 'peso_gramos', '-') if x['rn'] else '-'),
        ("RESULTADO DEL RN", "Talla", lambda x: safe_get(x['rn'], 'talla_centimetros', '-') if x['rn'] else '-'),
        ("RESULTADO DEL RN", "Ligadura Tardía del Cordón (> a 1 minuto)", lambda x: get_si_no(safe_get(x['rn'], 'ligadura_tardia_cordon', False)) if x['rn'] else '-'),
        ("RESULTADO DEL RN", "Apgar al minuto", lambda x: safe_get(x['rn'], 'apgar_1_minuto', '-') if x['rn'] else '-'),
        ("RESULTADO DEL RN", "Apgar a los 5 min", lambda x: safe_get(x['rn'], 'apgar_5_minutos', '-') if x['rn'] else '-'),
        ("RESULTADO DEL RN", "Tiempo de apego", lambda x: safe_get(x['rn'], 'tiempo_primer_apego_minutos', '-') if x['rn'] else '-'),

        # --- APEGO Y ACOMPAÑAMIENTO (New Image) ---
        ("APEGO Y ACOMPAÑAMIENTO", "Apego Canguro", lambda x: get_si_no(safe_get(x['rn'], 'apego_canguro', False)) if x['rn'] else '-'),
        ("APEGO Y ACOMPAÑAMIENTO", "Acompañamiento Preparto", lambda x: get_si_no(safe_get(x['parto'], 'acompanamiento_preparto', False))),
        ("APEGO Y ACOMPAÑAMIENTO", "Acompañamiento Parto", lambda x: get_si_no(safe_get(x['parto'], 'acompanamiento_parto', False))),
        ("APEGO Y ACOMPAÑAMIENTO", "Acomp. RN", lambda x: get_si_no(safe_get(x['parto'], 'acompanamiento_rn', False))),
        ("APEGO Y ACOMPAÑAMIENTO", "Motivo parto NO acompañado", lambda x: safe_get(x['parto'], 'motivo_parto_no_acompanado.descripcion', '-')),
        ("APEGO Y ACOMPAÑAMIENTO", "Persona acompañante", lambda x: safe_get(x['parto'], 'persona_acompanante.descripcion', '-')),
        ("APEGO Y ACOMPAÑAMIENTO", "Acompañante secciona cordón", lambda x: get_si_no(safe_get(x['parto'], 'acompanante_secciona_cordon', False))),

        # --- INFORMACION DE LOS PROFESIONALES (New Image) ---
        ("INFORMACION DE LOS PROFESIONALES", "Profes. Responsable (Nombre - apellido)", lambda x: f"{safe_get(x['parto'], 'profesional_responsable_nombre', '')} {safe_get(x['parto'], 'profesional_responsable_apellido', '')}".strip() or '-'),
        ("INFORMACION DE LOS PROFESIONALES", "Causa cesarea", lambda x: safe_get(x['parto'], 'causa_cesarea.descripcion', '-')),
        ("INFORMACION DE LOS PROFESIONALES", "OBSERVACIONES", lambda x: safe_get(x['parto'], 'observaciones', '-')),
        ("INFORMACION DE LOS PROFESIONALES", "USO DE SALA SAIP (SI/NO)", lambda x: get_si_no(safe_get(x['parto'], 'uso_sala_saip', False))),

        # --- LEY DOMINGA / PLACENTA ---
        ("Ley N° 21.372 Dominga", "Recuerdos Entregados", lambda x: safe_get(x['parto'], 'ley_dominga_recuerdos', '-')),
        ("Ley N° 21.372 Dominga", "Justificación (No entrega)", lambda x: safe_get(x['parto'], 'ley_dominga_justificacion', '-')),
        ("Placenta", "Retira Placenta", lambda x: get_si_no(safe_get(x['parto'], 'retira_placenta', False))),
        ("Placenta", "Estampado de Placenta", lambda x: get_si_no(safe_get(x['parto'], 'estampado_placenta', False))),
    ]
    
    # ---------------------------------------------------------
    # ESCRITURA DE ENCABEZADOS
    # ---------------------------------------------------------
    
    current_super_header = None
    start_merge_col = 1
    
    for col_num, (super_header, header, extractor) in enumerate(columns_map, 1):
        # 2. Fila Secundaria (Header Específico)
        cell_h = ws.cell(row=2, column=col_num, value=header)
        cell_h.font = header_font
        # Usar un color base para el header secundario (puedes variar si quieres)
        cell_h.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid") 
        cell_h.alignment = center_align
        cell_h.border = border_style

        # 1. Fila Principal (Super Header) con Merged Cells
        if super_header != current_super_header:
            if current_super_header is not None:
                ws.merge_cells(start_row=1, start_column=start_merge_col, end_row=1, end_column=col_num - 1)
                
            cell_s = ws.cell(row=1, column=col_num, value=super_header)
            cell_s.font = header_font
            
            # Asignar color según categoría
            color_hex = CATEGORY_COLORS.get(super_header, "000000")
            cell_s.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            cell_s.alignment = center_align
            cell_s.border = border_style

            current_super_header = super_header
            start_merge_col = col_num

    # Merge final
    if current_super_header is not None:
        ws.merge_cells(start_row=1, start_column=start_merge_col, end_row=1, end_column=len(columns_map))

    # ---------------------------------------------------------
    # OBTENCIÓN Y ESCRITURA DE DATOS CON FILTROS
    # ---------------------------------------------------------
    
    queryset = RegistroParto.objects.select_related(
        'ficha_obstetrica',
        'ficha_obstetrica__paciente',
        'ficha_obstetrica__paciente__persona',
        'ficha_obstetrica__tipo_paciente', # Include new relation
        'ficha_obstetrica__clasificacion_aro', # Include new relation
        'ficha_obstetrica__discapacidad', # Include new relation
        'ficha_obstetrica__paciente__persona__Nacionalidad',
        'ficha_obstetrica__paciente__persona__Pueblos_originarios',
        'ficha_obstetrica__consultorio_origen',
        'ficha_ingreso_parto',
        'tipo_parto',
        'clasificacion_robson',
        'posicion_parto',
        'estado_perine',
        'regimen_parto',
        'tipo_esterilizacion',
        'causa_cesarea',
        'persona_acompanante',
        'motivo_parto_no_acompanado'
    ).prefetch_related(
        'metodos_no_farmacologicos',
        'recien_nacidos',
        'recien_nacidos__sexo',
        'recien_nacidos__documentos_parto'
    )
    
    # APLICAR FILTROS DESDE REQUEST.POST
    from django.db.models import Q
    
    # Filtro por rango de fechas
    fecha_inicio = request.POST.get('fecha_inicio')
    fecha_fin = request.POST.get('fecha_fin')
    
    if fecha_inicio:
        queryset = queryset.filter(fecha_hora_parto__date__gte=fecha_inicio)
    if fecha_fin:
        queryset = queryset.filter(fecha_hora_parto__date__lte=fecha_fin)
    

    
    # Filtro por tipo de parto
    tipo_parto_id = request.POST.get('tipo_parto')
    if tipo_parto_id:
        queryset = queryset.filter(tipo_parto_id=tipo_parto_id)
    
    # Filtro por clasificación Robson
    robson_id = request.POST.get('clasificacion_robson')
    if robson_id:
        queryset = queryset.filter(clasificacion_robson_id=robson_id)
    
    # Ordenar
    queryset = queryset.order_by('-fecha_hora_parto')

    if not queryset.exists():
        ws.cell(row=3, column=1, value="NO SE ENCONTRARON REGISTROS")
    
    row_num = 3
    for parto in queryset:
        rns = list(parto.recien_nacidos.all())
        
        # Preparar lista de contextos y duplicar parto por cada RN
        contexts = []
        if rns:
            for rn in rns:
                contexts.append({'parto': parto, 'rn': rn})
        else:
            contexts.append({'parto': parto, 'rn': None})
            
        for ctx in contexts:
            for col_index, (sh, h, extractor) in enumerate(columns_map):
                try:
                    val = extractor(ctx)
                except Exception as e:
                    val = "-"
                
                if isinstance(val, str): val = val.strip()
                
                cell = ws.cell(row=row_num, column=col_index + 1, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_style
                
            row_num += 1

    # Ajustar ancho de columnas
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response
